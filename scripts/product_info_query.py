#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
商品信息查询脚本 - GDS API版本
根据Excel文件中的条码数字，调用中国商品信息服务平台（GDS）官方API查询商品信息
专门用于商品信息查询功能，需要提供条码数字列
"""

import argparse
import os
import shutil
import sys
import time
import requests
from openpyxl import load_workbook

def format_barcode(barcode_data):
    """
    格式化条码数据 - 如果条码长度为13位，则在首位补0
    此函数从 barcode_recognizer.py 迁移而来，确保查询前条码格式正确
    
    Args:
        barcode_data: 原始条码数据
    
    Returns:
        str: 格式化后的条码数据
    """
    if barcode_data and len(str(barcode_data).strip()) == 13:
        # 13位条码在首位补0，变为14位
        original_barcode = str(barcode_data).strip()
        formatted_barcode = '0' + original_barcode
        print(f"    条码格式化: {original_barcode} -> {formatted_barcode} (13位补0)")
        return formatted_barcode
    return str(barcode_data).strip() if barcode_data else ''

def parse_args():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(
        description='根据Excel文件中的条码数字查询商品信息（GDS API版本）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  python scripts/product_info_query.py /Users/bytedance/Desktop/tgs/条码识别结果_好客来超市-商品统计-latest.xlsx --barcode-column 5 --authorization-token YOUR_TOKEN
  python scripts/product_info_query.py /Users/bytedance/Desktop/tgs/条码识别结果_好客来超市-商品统计-latest.xlsx --barcode-column 5 --authorization-token YOUR_TOKEN --output result.xlsx --start-row 3
  
注意事项:
  - 需要先在中国商品信息服务平台（GDS）获取授权令牌
  - 使用官方API接口，数据更准确可靠
  - 商品信息将从Excel最后一列开始写入
  - 包含QPS限制，每秒最多1次请求
  - 每次查询完成后立即保存到文件，支持断点续传
  - 可指定起始行，默认从第2行开始（跳过标题行）
        """
    )
    
    parser.add_argument('excel_file', help='包含条码数字的Excel文件路径')
    parser.add_argument('--barcode-column', type=int, required=True,
                       help='包含条码数字的列号（从1开始计数）')
    parser.add_argument('--output', help='输出文件名（默认为原文件名加前缀）')
    parser.add_argument('--start-row', type=int, default=2,
                       help='开始处理的行号（默认为2，即跳过标题行）。\n'
                            '支持从指定行开始查询，便于断点续传或分批处理大量数据。')
    
    # GDS API配置
    parser.add_argument('--authorization-token', required=True, help='GDS API的授权令牌（Bearer Token）')
    parser.add_argument('--api-url', default="https://bff.gds.org.cn/gds/searching-api/ProductService/ProductListByGTIN",
                       help='GDS API地址（默认使用官方地址）')
    
    return parser.parse_args()

def query_product_info_gds(barcode, api_url, authorization_token, last_request_time=None):
    """
    使用中国商品信息服务平台（GDS）API查询商品信息
    
    Args:
        barcode: 条码数据
        api_url: GDS API地址
        authorization_token: 授权令牌
        last_request_time: 上次请求时间（用于QPS控制）
    
    Returns:
        dict: 包含查询结果的字典
    """
    # QPS限制：确保每次API请求间隔至少1秒
    # if last_request_time is not None:
    #     elapsed_time = time.time() - last_request_time
    #     if elapsed_time < 1.0:  # 如果距离上次请求不足1秒
    #         sleep_time = 1.0 - elapsed_time
    #         print(f"    QPS限制：等待 {sleep_time:.2f} 秒...")
    #         time.sleep(sleep_time)


    print(f"    QPS限制：等待 {3} 秒...")
    time.sleep(3)
    
    try:
        # 构建请求参数
        params = {
            'PageSize': 30,
            'PageIndex': 1,
            'SearchItem': barcode
        }
        
        # 构建请求头
        headers = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Authorization': f'Bearer {authorization_token}',
            'Connection': 'keep-alive',
            'Origin': 'https://www.gds.org.cn',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-site',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36',
            'currentRole': 'Mine',
            'sec-ch-ua': '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"macOS"'
        }
        
        # 发送API请求
        print(f"    正在查询商品信息: {barcode}")
        response = requests.get(api_url, params=params, headers=headers, timeout=1500)
        response.raise_for_status()
        
        # 解析JSON响应
        result = response.json()
        
        # 检查GDS API响应状态 - GDS API中Code=1表示成功
        if result.get('Code') == 1 and result.get('Data') and result['Data'].get('Items'):
            # 查询成功，获取第一个商品信息
            items = result['Data']['Items']
            if items and len(items) > 0:
                product = items[0]  # 取第一个匹配的商品
                
                # 构建返回数据，映射GDS字段到我们的标准字段
                result_data = {
                    'ProductName': product.get('RegulatedProductName', ''),  # 商品名称
                    'GTIN': product.get('gtin', barcode),  # 条码
                    'BrandName': product.get('brandcn', ''),  # 品牌
                    'CompanyName': product.get('firm_name', ''),  # 公司名称
                    'NetContent': product.get('specification', ''),  # 净含量/规格
                    'ProductDescription': product.get('description', '')  # 商品描述
                }
                
                return {
                    'success': True,
                    'data': result_data
                }
            else:
                return {
                    'success': False,
                    'error': '未找到匹配的商品信息'
                }
        else:
            # 查询失败，返回错误信息
            error_msg = result.get('Msg', '未知错误')
            return {
                'success': False,
                'error': f"GDS API返回错误: {error_msg}"
            }
            
    except requests.exceptions.RequestException as e:
        # 网络请求错误
        return {
            'success': False,
            'error': f"网络请求失败: {str(e)}"
        }
    except Exception as e:
        # 其他错误
        return {
            'success': False,
            'error': f"查询商品信息时发生错误: {str(e)}"
        }

def main():
    """主函数"""
    # 获取命令行参数
    args = parse_args()
    
    # 验证Excel文件是否存在
    if not os.path.exists(args.excel_file):
        print(f"错误: 文件 {args.excel_file} 不存在")
        return
    
    # 配置信息
    EXCEL_FILE = args.excel_file
    BARCODE_COLUMN = args.barcode_column
    API_URL = args.api_url
    AUTHORIZATION_TOKEN = args.authorization_token
    START_ROW = args.start_row  # 新增：起始行参数
    
    # 打印配置信息
    print(f"条码数字列: {BARCODE_COLUMN}")
    print(f"起始处理行: {START_ROW} (从此行开始查询条码)")
    print(f"GDS API地址: {API_URL}")
    print(f"授权令牌: {AUTHORIZATION_TOKEN[:20]}...")
    print(f"实时保存: 是（每次查询后立即保存到文件，确保数据不丢失）")
    print(f"处理模式: 逐行查询，支持断点续传和QPS限制")
    
    # 设置输出文件名
    if args.output:
        output_file = args.output
    else:
        output_file = f"product_商品信息查询结果_{os.path.basename(EXCEL_FILE)}"
    
    # 复制原始Excel文件
    print(f"创建Excel文件副本: {output_file}")
    shutil.copy2(EXCEL_FILE, output_file)
    
    # 加载工作簿（用于读取条码数据和写入结果）
    print(f"正在读取Excel文件中的条码数据: {EXCEL_FILE}")
    wb = load_workbook(output_file, read_only=False, keep_vba=True, data_only=False, keep_links=True)
    sheet = wb.active
    
    # 检测Excel的最后一列位置
    max_col = sheet.max_column
    start_col = max_col + 1  # 从最后一列的下一列开始写入商品信息
    
    # 定义GDS API返回的字段映射
    field_names = ['ProductName', 'GTIN', 'BrandName', 'CompanyName', 'NetContent', 'ProductDescription']
    field_headers = ['商品名称', '条码', '品牌', '公司名称', '净含量', '商品描述']
    
    print(f"将从第 {start_col} 列开始写入 {len(field_names)} 个字段")
    
    # 写入列标题（第1行）
    for i, header in enumerate(field_headers):
        sheet.cell(row=1, column=start_col + i, value=header)
    
    # QPS控制：记录上次API请求时间
    last_api_request_time = None
    
    # 获取Excel的最大行数
    max_row = sheet.max_row
    
    # 统计信息
    total_processed = 0
    success_count = 0
    
    # 计算需要处理的行数
    total_rows_to_process = max_row - START_ROW + 1
    print(f"\n将从第{START_ROW}行开始处理，共需处理 {total_rows_to_process} 行")
    
    # 处理每一行的条码数据（从指定起始行开始）
    for row in range(START_ROW, max_row + 1):
        try:
            # 显示处理进度
            current_progress = row - START_ROW + 1
            print(f"\n[{current_progress}/{total_rows_to_process}] 处理第{row}行...")
            
            # 读取条码数据
            barcode_cell = sheet.cell(row=row, column=BARCODE_COLUMN)
            barcode_data = barcode_cell.value
            
            # 跳过空白条码或无效条码
            # 这里处理各种无效条码情况，避免无效的API调用
            if not barcode_data or str(barcode_data).strip() == '' or str(barcode_data).strip() == '识别失败':
                print(f"  跳过空白或无效条码")
                # 写入跳过标记，便于后续统计和人工检查
                sheet.cell(row=row, column=start_col, value="跳过: 空白或无效条码")
                for i in range(1, len(field_names)):
                    sheet.cell(row=row, column=start_col + i, value='')
                # 立即保存（实时保存功能）
                wb.save(output_file)
                total_processed += 1
                continue
            
            # 转换为字符串并清理
            barcode_data = str(barcode_data).strip()
            
            # 格式化条码：13位条码补0处理
            # 这里实现了条码标准化，确保符合EAN-13格式要求
            formatted_barcode = format_barcode(barcode_data)
            
            print(f"  原始条码: {barcode_data}")
            if formatted_barcode != barcode_data:
                print(f"  格式化后条码: {formatted_barcode} (已补0至13位)")
            
            # 查询商品信息（使用GDS官方API，带QPS限制）
            product_result = query_product_info_gds(formatted_barcode, API_URL, AUTHORIZATION_TOKEN, last_api_request_time)
            
            # 更新上次API请求时间
            last_api_request_time = time.time()
            
            if product_result.get('success'):
                print(f"  ✓ 查询成功: {product_result['data'].get('ProductName', '未知商品')}")
                # 立即写入查询结果到Excel
                data = product_result['data']
                for i, field_name in enumerate(field_names):
                    value = data.get(field_name, '')
                    sheet.cell(row=row, column=start_col + i, value=value)
                success_count += 1
            else:
                print(f"  ❌ 查询失败: {product_result.get('error', '未知错误')}")
                # 写入错误信息到第一列，其他列留空
                error_msg = product_result.get('error', '未知错误')
                sheet.cell(row=row, column=start_col, value=f"错误: {error_msg}")
                for i in range(1, len(field_names)):
                    sheet.cell(row=row, column=start_col + i, value='')
            
            # 立即保存文件（实时保存功能）
            # 每次查询完成后立即保存，确保数据不会因程序中断而丢失
            wb.save(output_file)
            print(f"  💾 已保存到第{row}行")
            total_processed += 1
            
        except Exception as e:
            print(f"  ❌ 处理失败: {str(e)}")
            # 写入错误信息，便于后续排查问题
            sheet.cell(row=row, column=start_col, value=f"处理错误: {e}")
            for i in range(1, len(field_names)):
                sheet.cell(row=row, column=start_col + i, value='')
            # 立即保存（即使出错也要保存状态）
            wb.save(output_file)
            total_processed += 1
    
    # 显示最终统计信息
    print(f"\n🎉 所有条码查询完成！")
    print(f"📊 处理统计:")
    print(f"  - 总共处理: {total_processed} 个条码")
    print(f"  - 查询成功: {success_count} 个")
    print(f"  - 查询失败: {total_processed - success_count} 个")
    print(f"  - 成功率: {success_count/total_processed*100:.1f}%" if total_processed > 0 else "  - 成功率: 0%")
    print(f"💾 所有结果已实时保存到: {output_file}")
    print(f"\n✅ 任务完成！文件已保存，可以直接查看结果。")
    
    try:
        wb.save(output_file)
        print(f"\n注意: 使用GDS官方API进行商品信息查询")
        print(f"如果遇到API错误，请检查authorization-token是否正确或已过期")
        print(f"\n此脚本专门用于商品信息查询，如需条码识别请使用 barcode_recognizer.py 脚本")
    except Exception as e:
        print(f"保存文件时出错: {e}")
        # 尝试使用另一种方式保存
        try:
            print("尝试使用替代方法保存文件...")
            temp_output = f"temp_{output_file}"
            wb.save(temp_output)
            wb.close()
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_output, output_file)
            print(f"使用替代方法保存成功: {output_file}")
        except Exception as e2:
            print(f"使用替代方法保存失败: {e2}")

if __name__ == "__main__":
    main()