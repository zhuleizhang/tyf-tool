#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
商品条码识别脚本 - GDS API版本
从Excel文件中提取条码图片，识别条码并查询商品信息
使用中国商品信息服务平台（GDS）官方API进行商品信息查询
"""

import argparse
import os
import shutil
import sys
import time
from io import BytesIO
from PIL import Image, ImageEnhance, ImageFilter
import requests
from openpyxl import load_workbook
import numpy as np
try:
    import cv2
except ImportError:
    cv2 = None
    print("警告: OpenCV未安装，将使用基础图像处理方法")

# macOS环境变量配置 - 设置zbar库路径
def setup_macos_environment():
    """为macOS系统设置zbar库环境变量"""
    if sys.platform == 'darwin':  # macOS
        # 常见的zbar库安装路径
        possible_paths = [
            '/opt/homebrew/opt/zbar/lib',  # Apple Silicon Mac (M1/M2)
            '/usr/local/opt/zbar/lib',     # Intel Mac
            '/opt/local/lib',              # MacPorts
        ]
        
        for zbar_path in possible_paths:
            if os.path.exists(zbar_path):
                current_path = os.environ.get('DYLD_LIBRARY_PATH', '')
                if zbar_path not in current_path:
                    if current_path:
                        os.environ['DYLD_LIBRARY_PATH'] = f"{zbar_path}:{current_path}"
                    else:
                        os.environ['DYLD_LIBRARY_PATH'] = zbar_path
                print(f"已设置zbar库路径: {zbar_path}")
                return True
        
        print("警告: 未找到zbar库，请确保已正确安装")
        return False
    return True

# 检查和导入依赖库
def check_dependencies():
    """检查所有必要的依赖库及其版本"""
    missing_deps = []
    version_info = []
    
    # 检查pyzbar
    try:
        from pyzbar import pyzbar
        try:
            import pyzbar
            version = getattr(pyzbar, '__version__', '未知版本')
            version_info.append(f"pyzbar: {version}")
        except:
            version_info.append("pyzbar: 版本未知")
        print("✓ pyzbar库导入成功")
    except ImportError as e:
        missing_deps.append(('pyzbar', str(e)))
    
    # 检查其他依赖
    deps_to_check = [
        ('PIL', 'Pillow'),
        ('openpyxl', 'openpyxl'),
        ('requests', 'requests'),
        ('numpy', 'numpy')
    ]
    
    # 检查可选依赖OpenCV
    try:
        import cv2
        version = getattr(cv2, '__version__', '未知版本')
        version_info.append(f"opencv-python: {version}")
        print("✓ OpenCV库导入成功（用于高级图像处理）")
    except ImportError:
        print("⚠ OpenCV库未安装（可选，用于高级图像处理）")
    
    for module_name, package_name in deps_to_check:
        try:
            module = __import__(module_name)
            # 尝试获取版本信息
            version = 'unknown'
            for attr in ['__version__', 'version', 'VERSION']:
                if hasattr(module, attr):
                    version = getattr(module, attr)
                    if isinstance(version, tuple):
                        version = '.'.join(map(str, version))
                    break
            version_info.append(f"{package_name}: {version}")
            print(f"✓ {package_name}库导入成功")
        except ImportError as e:
            missing_deps.append((package_name, str(e)))
    
    # 显示版本信息
    if version_info:
        print("\n📦 依赖库版本信息:")
        for info in version_info:
            print(f"  {info}")
    
    if missing_deps:
        print("\n❌ 缺少以下依赖库:")
        for dep, error in missing_deps:
            print(f"  - {dep}: {error}")
        
        print("\n📋 解决方案:")
        print("1. 安装zbar系统库:")
        if sys.platform == 'darwin':
            print("   macOS: brew install zbar")
        elif sys.platform.startswith('linux'):
            print("   Ubuntu/Debian: sudo apt-get install libzbar0")
            print("   CentOS/RHEL: sudo yum install zbar")
        
        print("\n2. 安装Python依赖:")
        print("   pip install pyzbar openpyxl pillow requests numpy")
        print("   pip install opencv-python  # 可选，用于高级图像处理")
        
        print("\n3. 如果仍有问题，请尝试:")
        print("   pip install --upgrade pyzbar")
        print("   pip install --force-reinstall pyzbar")
        
        if sys.platform == 'darwin':
            print("\n4. macOS特殊情况:")
            print("   export DYLD_LIBRARY_PATH=/opt/homebrew/opt/zbar/lib:$DYLD_LIBRARY_PATH")
            print("   或者")
            print("   export DYLD_LIBRARY_PATH=/usr/local/opt/zbar/lib:$DYLD_LIBRARY_PATH")
        
        return False
    
    return True

# 设置环境并检查依赖
setup_macos_environment()
if not check_dependencies():
    print("\n❌ 依赖检查失败，请按照上述说明安装缺少的依赖库")
    sys.exit(1)

# 导入条码识别库
try:
    from pyzbar import pyzbar
except ImportError:
    print("❌ pyzbar导入失败，请检查安装")
    sys.exit(1)

def parse_args():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(
        description='从Excel文件中识别条码并查询商品信息（GDS API版本）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  python scripts/barcode_scanner_gf.py /Users/bytedance/Desktop/tgs/好客来超市-商品统计-latest.xlsx --image-cols 2 --authorization-token YOUR_TOKEN
  python scripts/barcode_scanner_gf.py /Users/bytedance/Desktop/tgs/好客来超市-商品统计.xlsx --image-cols 2 3 --authorization-token YOUR_TOKEN --output result.xlsx
  
注意事项:
  - 需要先在中国商品信息服务平台（GDS）获取授权令牌
  - 使用官方API接口，数据更准确可靠
  - 商品信息将从Excel最后一列开始写入
        """
    )
    
    parser.add_argument('excel_file', help='包含条码图片的Excel文件路径')
    parser.add_argument('--image-cols', type=int, nargs='+', required=True,
                       help='包含条码图片的列号（从1开始计数），可指定多列')
    parser.add_argument('--output', help='输出文件名（默认为原文件名加前缀）')
    
    # GDS API配置
    parser.add_argument('--authorization-token', required=True, help='GDS API的授权令牌（Bearer Token）')
    parser.add_argument('--api-url', default="https://bff.gds.org.cn/gds/searching-api/ProductService/ProductListByGTIN",
                       help='GDS API地址（默认使用官方地址）')
    
    return parser.parse_args()

def format_barcode(barcode_data):
    """
    格式化条码数据 - 如果条码长度为13位，则在首位补0
    
    Args:
        barcode_data: 原始条码数据
    
    Returns:
        str: 格式化后的条码数据
    """
    if barcode_data and len(barcode_data) == 13:
        # 13位条码在首位补0，变为14位
        formatted_barcode = '0' + barcode_data
        print(f"    条码格式化: {barcode_data} -> {formatted_barcode} (13位补0)")
        return formatted_barcode
    return barcode_data

def decode_barcode_from_image(image_data):
    """
    从图片数据中识别条码 - 增强版本
    针对圆柱体饮料条码等弯曲变形条码进行优化
    
    Args:
        image_data: 图片的二进制数据
    
    Returns:
        tuple: (条码数据, 条码类型) 或 (None, None)
    """
    try:
        # 将二进制数据转换为PIL图像对象
        original_image = Image.open(BytesIO(image_data))
        
        # 转换为RGB格式（确保兼容性）
        if original_image.mode != 'RGB':
            original_image = original_image.convert('RGB')
        
        print(f"    尝试识别条码，原图尺寸: {original_image.size}")
        
        # 策略1: 直接识别原图
        barcodes = pyzbar.decode(original_image)
        if barcodes:
            barcode = barcodes[0]
            barcode_data = barcode.data.decode('utf-8')
            formatted_barcode = format_barcode(barcode_data)
            print(f"    ✓ 原图识别成功: {formatted_barcode}")
            return formatted_barcode, barcode.type
        
        # 策略2: 基础图像预处理
        processed_images = []
        
        # 2.1 灰度化处理
        gray_image = original_image.convert('L')
        processed_images.append(("灰度化", gray_image))
        
        # 2.2 对比度增强
        enhancer = ImageEnhance.Contrast(original_image)
        contrast_image = enhancer.enhance(2.0)  # 增强对比度
        processed_images.append(("对比度增强", contrast_image))
        
        # 2.3 锐化处理
        sharp_image = original_image.filter(ImageFilter.SHARPEN)
        processed_images.append(("锐化处理", sharp_image))
        
        # 2.4 高斯模糊去噪
        blur_image = original_image.filter(ImageFilter.GaussianBlur(radius=0.5))
        processed_images.append(("高斯模糊", blur_image))
        
        # 2.5 亮度调整
        brightness_enhancer = ImageEnhance.Brightness(original_image)
        bright_image = brightness_enhancer.enhance(1.2)
        processed_images.append(("亮度增强", bright_image))
        
        # 尝试识别预处理后的图像
        for method_name, processed_image in processed_images:
            barcodes = pyzbar.decode(processed_image)
            if barcodes:
                barcode = barcodes[0]
                barcode_data = barcode.data.decode('utf-8')
                formatted_barcode = format_barcode(barcode_data)
                print(f"    ✓ {method_name}识别成功: {formatted_barcode}")
                return formatted_barcode, barcode.type
        
        # 策略3: 多角度旋转识别
        print("    尝试多角度旋转识别...")
        for angle in range(-10, 11, 2):  # -10度到+10度，步长2度
            if angle == 0:  # 0度已经在原图中尝试过了
                continue
            
            rotated_image = original_image.rotate(angle, expand=True)
            barcodes = pyzbar.decode(rotated_image)
            if barcodes:
                barcode = barcodes[0]
                barcode_data = barcode.data.decode('utf-8')
                formatted_barcode = format_barcode(barcode_data)
                print(f"    ✓ 旋转{angle}度识别成功: {formatted_barcode}")
                return formatted_barcode, barcode.type
        
        # 策略4: 缩放识别
        print("    尝试缩放识别...")
        for scale in [0.8, 1.2, 1.5]:  # 不同缩放比例
            width, height = original_image.size
            new_size = (int(width * scale), int(height * scale))
            scaled_image = original_image.resize(new_size, Image.Resampling.LANCZOS)
            
            barcodes = pyzbar.decode(scaled_image)
            if barcodes:
                barcode = barcodes[0]
                barcode_data = barcode.data.decode('utf-8')
                formatted_barcode = format_barcode(barcode_data)
                print(f"    ✓ 缩放{scale}x识别成功: {formatted_barcode}")
                return formatted_barcode, barcode.type
        
        # 策略5: 裁剪中心区域识别
        print("    尝试裁剪中心区域识别...")
        width, height = original_image.size
        # 裁剪中心80%的区域
        crop_margin_w = int(width * 0.1)
        crop_margin_h = int(height * 0.1)
        cropped_image = original_image.crop((
            crop_margin_w, crop_margin_h, 
            width - crop_margin_w, height - crop_margin_h
        ))
        
        barcodes = pyzbar.decode(cropped_image)
        if barcodes:
            barcode = barcodes[0]
            barcode_data = barcode.data.decode('utf-8')
            formatted_barcode = format_barcode(barcode_data)
            print(f"    ✓ 裁剪中心区域识别成功: {formatted_barcode}")
            return formatted_barcode, barcode.type
        
        # 策略6: 使用OpenCV进行高级处理（如果可用）
        if cv2 is not None:
            print("    尝试OpenCV高级处理...")
            # 转换为OpenCV格式
            cv_image = cv2.cvtColor(np.array(original_image), cv2.COLOR_RGB2BGR)
            
            # 6.1 自适应二值化
            gray = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
            adaptive_thresh = cv2.adaptiveThreshold(
                gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2
            )
            adaptive_image = Image.fromarray(adaptive_thresh)
            
            barcodes = pyzbar.decode(adaptive_image)
            if barcodes:
                barcode = barcodes[0]
                barcode_data = barcode.data.decode('utf-8')
                formatted_barcode = format_barcode(barcode_data)
                print(f"    ✓ 自适应二值化识别成功: {formatted_barcode}")
                return formatted_barcode, barcode.type
            
            # 6.2 形态学操作
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
            morph_image = cv2.morphologyEx(gray, cv2.MORPH_CLOSE, kernel)
            morph_pil = Image.fromarray(morph_image)
            
            barcodes = pyzbar.decode(morph_pil)
            if barcodes:
                barcode = barcodes[0]
                barcode_data = barcode.data.decode('utf-8')
                formatted_barcode = format_barcode(barcode_data)
                print(f"    ✓ 形态学处理识别成功: {formatted_barcode}")
                return formatted_barcode, barcode.type
        
        # 所有策略都失败
        print("    ✗ 所有识别策略均失败")
        return None, None
            
    except Exception as e:
        print(f"    条码识别错误: {e}")
        return None, None

def query_product_info_gds(barcode, api_url, authorization_token, last_request_time=None):
    """
    使用中国商品信息服务平台（GDS）API查询商品信息
    
    Args:
        barcode: 条码数据
        api_url: GDS API地址
        authorization_token: 授权令牌
        last_request_time: 上次请求时间（用于QPS控制）
    
    Returns:
        dict: 包含查询结果的字典
    """
    # QPS限制：确保每次API请求间隔至少1秒
    if last_request_time is not None:
        elapsed_time = time.time() - last_request_time
        if elapsed_time < 1.0:  # 如果距离上次请求不足1秒
            sleep_time = 1.0 - elapsed_time
            print(f"    QPS限制：等待 {sleep_time:.2f} 秒...")
            time.sleep(sleep_time)
    try:
        # 构建请求参数
        params = {
            'PageSize': 30,
            'PageIndex': 1,
            'SearchItem': barcode
        }
        
        # 构建请求头
        headers = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Authorization': f'Bearer {authorization_token}',
            'Connection': 'keep-alive',
            'Origin': 'https://www.gds.org.cn',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-site',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36',
            'currentRole': 'Mine',
            'sec-ch-ua': '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"macOS"'
        }
        
        # 发送API请求
        print(f"    正在查询商品信息: {barcode}")
        response = requests.get(api_url, params=params, headers=headers, timeout=15)
        response.raise_for_status()
        
        # 解析JSON响应
        result = response.json()
        
        # 检查GDS API响应状态 - GDS API中Code=1表示成功
        if result.get('Code') == 1 and result.get('Data') and result['Data'].get('Items'):
            # 查询成功，获取第一个商品信息
            items = result['Data']['Items']
            if items and len(items) > 0:
                product = items[0]  # 取第一个匹配的商品
                
                # 构建返回数据，映射GDS字段到我们的标准字段
                result_data = {
                    'ProductName': product.get('RegulatedProductName', ''),  # 商品名称
                    'GTIN': product.get('gtin', barcode),  # 条码
                    'BrandName': product.get('brandcn', ''),  # 品牌
                    'CompanyName': product.get('firm_name', ''),  # 公司名称
                    'NetContent': product.get('specification', ''),  # 净含量/规格
                    'ProductDescription': product.get('description', '')  # 商品描述
                }
                
                return {
                    'success': True,
                    'data': result_data
                }
            else:
                return {
                    'success': False,
                    'error': '未找到匹配的商品信息'
                }
        else:
            # 查询失败，返回错误信息
            error_msg = result.get('Msg', '未知错误')
            return {
                'success': False,
                'error': f"GDS API返回错误: {error_msg}"
            }
            
    except requests.exceptions.RequestException as e:
        # 网络请求错误
        return {
            'success': False,
            'error': f"网络请求失败: {str(e)}"
        }
    except Exception as e:
        # 其他错误
        return {
            'success': False,
            'error': f"查询商品信息时发生错误: {str(e)}"
        }

def main():
    """主函数"""
    # 获取命令行参数
    args = parse_args()
    
    # 验证Excel文件是否存在
    if not os.path.exists(args.excel_file):
        print(f"错误: 文件 {args.excel_file} 不存在")
        return
    
    # 配置信息
    EXCEL_FILE = args.excel_file
    IMAGE_COLUMNS = args.image_cols
    API_URL = args.api_url
    AUTHORIZATION_TOKEN = args.authorization_token
    
    # 打印配置信息
    print(f"条码图片列: {IMAGE_COLUMNS}")
    print(f"GDS API地址: {API_URL}")
    print(f"授权令牌: {AUTHORIZATION_TOKEN[:20]}...")
    
    # 设置输出文件名
    if args.output:
        output_file = args.output
    else:
        output_file = f"gds_条码查询结果_{os.path.basename(EXCEL_FILE)}"
    
    # 复制原始Excel文件
    print(f"创建Excel文件副本: {output_file}")
    shutil.copy2(EXCEL_FILE, output_file)
    
    # 加载原始工作簿（用于读取图片数据）
    print(f"正在读取原始Excel文件中的条码图片: {EXCEL_FILE}")
    source_wb = load_workbook(EXCEL_FILE)
    source_sheet = source_wb.active
    
    # 检测Excel的最后一列位置
    max_col = source_sheet.max_column
    start_col = max_col + 1  # 从最后一列的下一列开始写入
    
    # 定义GDS API返回的字段映射
    field_names = ['ProductName', 'GTIN', 'BrandName', 'CompanyName', 'NetContent', 'ProductDescription']
    field_headers = ['商品名称', '条码', '品牌', '公司名称', '净含量', '商品描述']
    
    print(f"将从第 {start_col} 列开始写入 {len(field_names)} 个字段")
    
    # 收集条码识别和商品查询结果
    query_results = {}  # 格式: {行号: 商品信息结构化数据}
    image_positions = []
    
    # QPS控制：记录上次API请求时间
    last_api_request_time = None
    
    # 获取所有图片的位置信息
    for idx, img in enumerate(source_sheet._images):
        row = img.anchor._from.row + 1  # 转换为1-indexed
        col = img.anchor._from.col + 1  # 转换为1-indexed
        image_positions.append((idx, row, col))
    
    # 处理每个图片列
    for image_col in IMAGE_COLUMNS:
        print(f"\n处理条码图片列 {image_col}")
        
        # 找出当前图片列中的所有图片
        column_images = [(idx, row, col) for idx, row, col in image_positions if col == image_col]
        
        if not column_images:
            print(f"警告: 列 {image_col} 中未找到图片")
            continue
        
        # 处理当前列的每个图片
        for idx, row, col in column_images:
            try:
                # 获取图片数据
                img = source_sheet._images[idx]
                img_data = img._data()

                print(f" ============= 行 {row}: 开始识别 ============= \n")
                
                # 识别条码
                barcode_data, barcode_type = decode_barcode_from_image(img_data)
                
                if barcode_data:
                    print(f"  行 {row}: 识别到条码 {barcode_data} (类型: {barcode_type})")
                    
                    # 查询商品信息（使用GDS官方API，带QPS限制）
                    product_result = query_product_info_gds(barcode_data, API_URL, AUTHORIZATION_TOKEN, last_api_request_time)
                    
                    # 更新上次API请求时间
                    last_api_request_time = time.time()
                    
                    # 如果商品信息查询失败，但条码识别成功，创建包含条码信息的结果结构
                    if not product_result.get('success'):
                        print(f"  查询失败: {product_result.get('error', '未知错误')}")
                        print(f"  条码识别成功，商品信息查询失败，仅填入条码")
                        # 创建包含条码信息的结果结构，其他字段为空
                        query_results[row] = {
                            'success': True,  # 标记为成功，因为条码识别成功
                            'barcode_only': True,  # 标记这是仅有条码的情况
                            'data': {
                                'ProductName': '',
                                'GTIN': barcode_data,  # 填入识别到的条码
                                'BrandName': '',
                                'CompanyName': '',
                                'NetContent': '',
                                'ProductDescription': ''
                            }
                        }
                    else:
                        print(f"  商品信息查询成功")
                        # 保存完整的查询结果
                        query_results[row] = product_result
                    
                else:
                    print(f"  行 {row}: 未识别到条码")
                    query_results[row] = {'success': False, 'error': '未识别到条码'}
                
                print(f" ============= 行 {row}: 识别结束 ============= \n")
            except Exception as e:
                print(f"  处理行 {row} 图片时出错: {e}")
                print(f" ============= 行 {row}: 识别失败 ============= \n")
                query_results[row] = {'success': False, 'error': f'处理错误: {e}'}
    
    # 关闭原始工作簿
    source_wb.close()
    
    # 加载复制后的工作簿（用于写入结果）
    print(f"\n正在将商品信息写入到: {output_file}")
    target_wb = load_workbook(output_file, read_only=False, keep_vba=True, data_only=False, keep_links=True)
    target_sheet = target_wb.active
    
    # 写入列标题（第1行）
    for i, header in enumerate(field_headers):
        target_sheet.cell(row=1, column=start_col + i, value=header)
    
    # 将查询结果写入多列
    for row, result in query_results.items():
        if result.get('success') and result.get('data'):
            # 成功获取商品信息或条码识别成功（包括仅有条码的情况），按字段写入各列
            data = result['data']
            for i, field_name in enumerate(field_names):
                value = data.get(field_name, '')
                target_sheet.cell(row=row, column=start_col + i, value=value)
            
            # 如果是仅有条码的情况，在日志中记录
            if result.get('barcode_only'):
                print(f"  行 {row}: 已填入条码 {data.get('barcode', '')}，其他信息为空")
        else:
            # 条码识别失败，在第一列写入错误信息
            error_msg = result.get('error', '未知错误')
            target_sheet.cell(row=row, column=start_col, value=f"错误: {error_msg}")
            # 其他列留空
            for i in range(1, len(field_names)):
                target_sheet.cell(row=row, column=start_col + i, value='')
    
    # 保存结果
    try:
        target_wb.save(output_file)
        print(f"处理完成，结果已保存到 {output_file}")
        print(f"\n共处理 {len(query_results)} 个条码图片")
        
        # 统计不同类型的结果
        full_success_count = sum(1 for result in query_results.values() 
                               if result.get('success') and not result.get('barcode_only'))
        barcode_only_count = sum(1 for result in query_results.values() 
                               if result.get('success') and result.get('barcode_only'))
        failed_count = sum(1 for result in query_results.values() if not result.get('success'))
        
        print(f"成功查询到完整商品信息: {full_success_count} 个")
        print(f"条码识别成功但商品信息查询失败: {barcode_only_count} 个")
        print(f"条码识别失败: {failed_count} 个")
        print(f"\n注意: 使用GDS官方API进行商品信息查询")
        print(f"如果遇到API错误，请检查authorization-token是否正确或已过期")
    except Exception as e:
        print(f"保存文件时出错: {e}")
        # 尝试使用另一种方式保存
        try:
            print("尝试使用替代方法保存文件...")
            temp_output = f"temp_{output_file}"
            target_wb.save(temp_output)
            target_wb.close()
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_output, output_file)
            print(f"使用替代方法保存成功: {output_file}")
        except Exception as e2:
            print(f"使用替代方法保存失败: {e2}")

if __name__ == "__main__":
    main()