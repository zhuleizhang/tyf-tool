#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
阿里云市场条码查询脚本

功能说明：
- 从Excel文件的指定条码数字列读取条码
- 通过阿里云市场条码查询API查询商品信息
- 提取ItemName、gpcname、ItemClassName三个字段
- 将查询结果追加到Excel文件的最后空列
- 支持实时保存和断点续传
- 提供详细的统计信息

使用示例：
python scripts/barcode_scanner_alicloud.py /Users/bytedance/Desktop/tgs/条码识别结果_好客来超市-商品统计-latest-第二次.xlsx --barcode-cols 5 --appcode your_appcode
python scripts/barcode_scanner_alicloud.py /Users/bytedance/Desktop/tgs/条码识别结果_好客来超市-商品统计-latest-第二次.xlsx --barcode-cols 5 --appcode your_appcode --output custom_output.xlsx --start-row 3

作者：SOLO Coding
版本：1.0
"""

import argparse
import json
import os
import re
import shutil
import sys
import time
import urllib3
from datetime import datetime
from openpyxl import load_workbook
from typing import Dict, Optional, Tuple

# 禁用urllib3的SSL警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# 阿里云市场条码查询API配置
API_HOST = 'https://barcode100.market.alicloudapi.com'
API_PATH = '/getBarcode'

def validate_barcode(barcode: str) -> bool:
    """
    验证条码格式
    
    Args:
        barcode: 条码字符串
        
    Returns:
        bool: 条码格式是否有效
    """
    if not barcode or not isinstance(barcode, str):
        return False
    
    # 移除空格和特殊字符
    barcode = re.sub(r'[^0-9]', '', str(barcode))
    
    # 检查是否为纯数字且长度合理（通常8-14位）
    if not barcode.isdigit() or len(barcode) < 8 or len(barcode) > 14:
        return False
        
    return True

def query_product_info_alicloud(barcode: str, appcode: str, http_pool: urllib3.PoolManager) -> Dict:
    """
    通过阿里云市场API查询商品信息
    
    Args:
        barcode: 条码
        appcode: 阿里云AppCode
        http_pool: HTTP连接池
        
    Returns:
        Dict: 查询结果，包含商品信息或错误信息
    """
    try:
        # 构建请求URL
        url = f"{API_HOST}{API_PATH}?Code={barcode}"
        
        # 设置请求头
        headers = {
            'Authorization': f'APPCODE {appcode}',
            'Content-Type': 'application/json'
        }
        
        # 发送请求
        response = http_pool.request('GET', url, headers=headers, timeout=10)
        
        # 检查HTTP状态码
        if response.status != 200:
            return {
                'success': False,
                'error': f'HTTP错误: {response.status}',
                'barcode': barcode
            }
        
        # 解析响应内容
        content = response.data.decode('utf-8')
        if not content:
            return {
                'success': False,
                'error': '响应内容为空',
                'barcode': barcode
            }
        
        # 解析JSON
        try:
            data = json.loads(content)
        except json.JSONDecodeError as e:
            return {
                'success': False,
                'error': f'JSON解析错误: {str(e)}',
                'barcode': barcode
            }
        
        # 检查API响应状态
        if data.get('status') != '200':
            return {
                'success': False,
                'error': f"API错误: {data.get('message', '未知错误')}",
                'barcode': barcode
            }
        
        # 提取所需字段
        result = {
            'success': True,
            'barcode': barcode,
            'ItemName': data.get('ItemName', ''),
            'gpcname': data.get('gpcname', ''),
            'ItemClassName': data.get('ItemClassName', '')
        }
        
        return result
        
    except Exception as e:
        return {
            'success': False,
            'error': f'请求异常: {str(e)}',
            'barcode': barcode
        }

def get_barcode_from_cell(cell_value) -> Optional[str]:
    """
    从Excel单元格中提取条码
    
    Args:
        cell_value: 单元格值
        
    Returns:
        Optional[str]: 提取的条码，如果无效则返回None
    """
    if cell_value is None:
        return None
    
    # 转换为字符串并清理
    barcode_str = str(cell_value).strip()
    
    # 移除可能的前缀（如"条码："）
    barcode_str = re.sub(r'^[^0-9]*', '', barcode_str)
    
    # 只保留数字
    barcode_str = re.sub(r'[^0-9]', '', barcode_str)
    
    if validate_barcode(barcode_str):
        return barcode_str
    
    return None

def write_results_to_file(original_file: str, output_file: str, results: list, new_columns: list, start_row: int):
    """
    将查询结果一次性写入Excel文件
    
    Args:
        original_file: 原始Excel文件路径
        output_file: 输出Excel文件路径
        results: 查询结果列表
        new_columns: 新增列信息
        start_row: 开始行号
    """
    try:
        # 复制原文件到输出文件
        shutil.copy2(original_file, output_file)
        print(f"已复制原文件到: {output_file}")
        
        # 加载工作簿
        workbook = load_workbook(output_file)
        worksheet = workbook.active
        
        # 检测Excel文件的最后一列
        last_col = worksheet.max_column
        print(f"检测到Excel文件最后一列: {last_col}")
        
        # 写入列标题
        for title, col_idx in new_columns:
            worksheet.cell(row=1, column=col_idx, value=title)
        
        print(f"已添加新列标题: {[title for title, _ in new_columns]}")
        
        # 写入查询结果
        for result in results:
            row_idx = result['row_idx']
            worksheet.cell(row=row_idx, column=last_col + 1, value=result.get('ItemName', ''))
            worksheet.cell(row=row_idx, column=last_col + 2, value=result.get('ItemClassName', ''))
            worksheet.cell(row=row_idx, column=last_col + 3, value=result.get('gpcname', ''))
        
        # 一次性保存文件
        workbook.save(output_file)
        workbook.close()
        print(f"已成功保存所有结果到: {output_file}")
        
    except Exception as e:
        print(f"写入文件时出错: {str(e)}")
        raise

def main():
    """
    主函数
    """
    # 解析命令行参数
    parser = argparse.ArgumentParser(
        description='阿里云市场条码查询脚本',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  python scripts/barcode_scanner_alicloud.py input.xlsx --barcode-cols 5 --appcode your_appcode
  python scripts/barcode_scanner_alicloud.py input.xlsx --barcode-cols 5 --appcode your_appcode --output result.xlsx --start-row 3

功能特点:
  - 从Excel条码数字列读取条码进行查询
  - 提取ItemName、gpcname、ItemClassName三个字段
  - 内存累积数据，一次性写入文件
  - 详细的错误处理和统计信息
  - 输出文件前缀: alicloud_条码查询结果_
        """
    )
    
    parser.add_argument('excel_file', help='Excel文件路径')
    parser.add_argument('--barcode-cols', type=int, required=True,
                       help='条码数字列号（从1开始计数）')
    parser.add_argument('--appcode', required=True,
                       help='阿里云市场AppCode')
    parser.add_argument('--output', 
                       help='输出文件名（可选，默认自动生成）')
    parser.add_argument('--start-row', type=int, default=2,
                       help='开始处理的行号（默认从第2行开始，跳过标题行）')
    
    args = parser.parse_args()
    
    # 验证输入文件
    if not os.path.exists(args.excel_file):
        print(f"错误：文件不存在 - {args.excel_file}")
        sys.exit(1)
    
    if not args.excel_file.lower().endswith(('.xlsx', '.xls')):
        print("错误：请提供Excel文件（.xlsx或.xls格式）")
        sys.exit(1)
    
    # 验证列号
    if args.barcode_cols < 1:
        print("错误：列号必须大于0")
        sys.exit(1)
    
    # 验证起始行号
    if args.start_row < 1:
        print("错误：起始行号必须大于0")
        sys.exit(1)
    
    # 生成输出文件名
    if args.output:
        output_file = args.output
    else:
        base_name = os.path.splitext(os.path.basename(args.excel_file))[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"alicloud_条码查询结果_{base_name}_{timestamp}.xlsx"
    
    print(f"\n=== 阿里云市场条码查询脚本 ===")
    print(f"输入文件: {args.excel_file}")
    print(f"条码列: 第{args.barcode_cols}列")
    print(f"起始行: 第{args.start_row}行")
    print(f"输出文件: {output_file}")
    print(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*50)
    
    try:
        # 临时加载原文件以获取基本信息
        temp_workbook = load_workbook(args.excel_file)
        temp_worksheet = temp_workbook.active
        
        # 检测Excel文件的最后一列和最大行数
        last_col = temp_worksheet.max_column
        max_row = temp_worksheet.max_row
        print(f"检测到Excel文件最后一列: {last_col}，最大行数: {max_row}")
        
        # 定义新增列的标题和位置
        new_columns = [
            ('商品名称', last_col + 1),
            ('商品分类名称', last_col + 2), 
            ('GPC分类名称', last_col + 3)
        ]
        
        # 关闭临时工作簿
        temp_workbook.close()
        
        print(f"将添加新列标题: {[title for title, _ in new_columns]}")
        
        # 创建HTTP连接池
        http_pool = urllib3.PoolManager(
            cert_reqs='CERT_NONE',
            ca_certs=None,
            retries=urllib3.Retry(total=3, backoff_factor=0.3)
        )
        
        # 统计变量
        total_processed = 0
        success_count = 0
        error_count = 0
        empty_count = 0
        
        # 内存中的查询结果列表
        query_results = []
        
        # 获取总行数
        total_rows = max(0, max_row - args.start_row + 1)
        
        print(f"\n开始处理条码查询，共{total_rows}行数据...")
        print(f"注意：所有数据将在内存中累积，完成后一次性写入文件")
        print("-" * 50)
        
        try:
            # 临时加载原文件以读取条码数据
            temp_workbook = load_workbook(args.excel_file)
            temp_worksheet = temp_workbook.active
            
            # 逐行处理
            for row_idx in range(args.start_row, max_row + 1):
                try:
                    # 获取条码
                    barcode_cell = temp_worksheet.cell(row=row_idx, column=args.barcode_cols)
                    barcode = get_barcode_from_cell(barcode_cell.value)
                    
                    if not barcode:
                        print(f"第{row_idx}行: 条码为空或格式无效，跳过")
                        empty_count += 1
                        total_processed += 1
                        continue
                    
                    print(f"第{row_idx}行: 查询条码 {barcode}...", end=" ")
                    
                    # 查询商品信息
                    result = query_product_info_alicloud(barcode, args.appcode, http_pool)
                    
                    # 准备结果数据
                    result_data = {
                        'row_idx': row_idx,
                        'barcode': barcode
                    }
                    
                    if result['success']:
                        # 查询成功，保存商品信息
                        result_data.update({
                            'ItemName': result.get('ItemName', ''),
                            'ItemClassName': result.get('ItemClassName', ''),
                            'gpcname': result.get('gpcname', '')
                        })
                        
                        success_count += 1
                        print(f"✓ 成功 - {result.get('ItemName', '未知商品')}")
                    else:
                        # 查询失败，只填入条码信息
                        result_data.update({
                            'ItemName': f"查询失败({barcode})",
                            'ItemClassName': '',
                            'gpcname': ''
                        })
                        
                        error_count += 1
                        print(f"✗ 失败 - {result.get('error', '未知错误')}")
                    
                    # 将结果添加到内存列表
                    query_results.append(result_data)
                    total_processed += 1
                    
                    # 显示进度
                    progress = (total_processed / total_rows) * 100
                    print(f"    进度: {total_processed}/{total_rows} ({progress:.1f}%) - 已查询: {len(query_results)}条")
                    
                    # API限流，避免请求过快
                    time.sleep(0.1)
                    
                except Exception as e:
                    print(f"第{row_idx}行处理异常: {str(e)}")
                    error_count += 1
                    total_processed += 1
                    continue
            
            # 关闭临时工作簿
            temp_workbook.close()
            
            # 所有数据处理完成，一次性写入文件
            print(f"\n数据查询完成，开始写入文件...")
            write_results_to_file(args.excel_file, output_file, query_results, new_columns, args.start_row)
                        
        except KeyboardInterrupt:
            print(f"\n用户中断操作，正在保存已查询的数据...")
            if query_results:
                try:
                    interrupted_file = output_file.replace('.xlsx', '_interrupted.xlsx')
                    write_results_to_file(args.excel_file, interrupted_file, query_results, new_columns, args.start_row)
                    print(f"已保存{len(query_results)}条查询结果到: {interrupted_file}")
                except Exception as interrupt_save_error:
                    print(f"中断保存失败: {str(interrupt_save_error)}")
            else:
                print(f"没有查询结果需要保存")
            raise
        
        # 输出统计信息
        print("\n" + "="*50)
        print("处理完成！统计信息：")
        print(f"总处理行数: {total_processed}")
        print(f"查询成功: {success_count}")
        print(f"查询失败: {error_count}")
        print(f"空条码: {empty_count}")
        print(f"成功率: {(success_count/max(1, total_processed-empty_count)*100):.1f}%")
        print(f"已保存结果: {len(query_results)}条")
        print(f"输出文件: {output_file}")
        print(f"结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("="*50)
        
    except Exception as e:
        print(f"\n程序执行出错: {str(e)}")
        # 尝试保存当前查询结果到备份文件
        if 'query_results' in locals() and query_results:
            try:
                backup_file = output_file.replace('.xlsx', '_error_backup.xlsx')
                write_results_to_file(args.excel_file, backup_file, query_results, new_columns, args.start_row)
                print(f"已保存{len(query_results)}条查询结果到错误备份文件: {backup_file}")
            except Exception as backup_error:
                print(f"错误备份保存失败: {str(backup_error)}")
        sys.exit(1)
    finally:
        # 清理资源
        if 'http_pool' in locals():
            http_pool.clear()
            print("HTTP连接池已清理")

if __name__ == '__main__':
    main()