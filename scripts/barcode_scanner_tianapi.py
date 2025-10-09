#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
商品条码查询脚本 (改造版)

功能说明:
- 从Excel文件的指定列中读取条码数字（不再需要图片识别）
- 自动验证条码格式（8-14位纯数字）
- 使用天聚数行API查询商品详细信息
- 实时写入查询结果到Excel文件，支持断点续传
- 支持多列条码数据批量处理

改造说明:
本脚本已从图片条码识别功能改造为直接处理条码数字的查询功能，
移除了所有图像处理相关的依赖和代码，专注于条码数据的API查询。

使用方法:
python barcode_scanner_tianapi.py --excel-file 数据文件.xlsx --barcode-cols 5 --tianapi-key YOUR_API_KEY --start-row 2

参数说明:
--excel-file: Excel文件路径
--barcode-cols: 包含条码数字的列号（从1开始，多列用逗号分隔）
--tianapi-key: 天聚数行API密钥
--start-row: 开始处理的行号（默认为2，跳过标题行）
--output: 输出文件名（可选，默认自动生成）

依赖安装:
pip install openpyxl requests

注意事项:
1. 条码数字必须为8-14位的纯数字格式
2. 每次查询后立即保存Excel文件，确保数据不丢失
3. 支持多列条码数据同时处理
4. 查询结果包含商品名称、规格、品牌、厂商等13个字段
5. 失败的查询会在Excel中标记错误信息，便于后续处理
"""

import argparse
import os
import shutil
import sys
import re
import requests
from openpyxl import load_workbook

# 条码格式验证函数
def validate_barcode(barcode_str):
    """
    验证条码格式是否有效
    
    Args:
        barcode_str: 条码字符串
    
    Returns:
        tuple: (是否有效, 清理后的条码)
    """
    if not barcode_str:
        return False, None
    
    # 转换为字符串并去除空白字符
    barcode_clean = str(barcode_str).strip()
    
    # 检查是否为纯数字
    if not re.match(r'^\d+$', barcode_clean):
        return False, None
    
    # 检查长度是否合理（一般条码长度在8-14位之间）
    if len(barcode_clean) < 8 or len(barcode_clean) > 14:
        return False, None
    
    return True, barcode_clean

# 检查基础依赖库
def check_dependencies():
    """检查必要的依赖库"""
    missing_deps = []
    
    # 检查基础依赖
    deps_to_check = [
        ('openpyxl', 'openpyxl'),
        ('requests', 'requests')
    ]
    
    for module_name, package_name in deps_to_check:
        try:
            __import__(module_name)
            print(f"✓ {package_name}库导入成功")
        except ImportError as e:
            missing_deps.append((package_name, str(e)))
    
    if missing_deps:
        print("\n❌ 缺少以下依赖库:")
        for dep, error in missing_deps:
            print(f"  - {dep}: {error}")
        
        print("\n📋 解决方案:")
        print("安装Python依赖:")
        print("   pip install openpyxl requests")
        
        return False
    
    return True

# 检查依赖
if not check_dependencies():
    print("\n❌ 依赖检查失败，请按照上述说明安装缺少的依赖库")
    sys.exit(1)

def parse_args():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(
        description='从Excel文件中读取条码数字并查询商品信息（天聚数行API版本）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  python scripts/barcode_scanner_tianapi.py /Users/bytedance/Desktop/tgs/条码识别结果_好客来超市-商品统计-latest.xlsx --barcode-cols 5 --tianapi-key YOUR_API_KEY
  python scripts/barcode_scanner_tianapi.py /Users/bytedance/Desktop/tgs/条码识别结果_好客来超市-商品统计-latest.xlsx --barcode-cols 5 --tianapi-key YOUR_API_KEY --output result.xlsx
  python scripts/barcode_scanner_tianapi.py /Users/bytedance/Desktop/tgs/条码识别结果_好客来超市-商品统计-latest.xlsx --barcode-cols 5 --tianapi-key YOUR_API_KEY --start-row 5

注意事项:
  - 需要先申请天聚数行API密钥（提供免费测试额度）
  - 商品信息将从Excel最后一列开始写入
  - 条码数字必须是8-14位的纯数字格式
  - 支持指定起始行，默认从第2行开始处理
  - 每次查询完成后立即保存到文件
  - 支持断点续传，可从指定行开始处理
        """
    )
    
    parser.add_argument('excel_file', help='包含条码数字的Excel文件路径')
    parser.add_argument('--barcode-cols', type=int, nargs='+', required=True,
                       help='包含条码数字的列号（从1开始计数），可指定多列')
    parser.add_argument('--output', help='输出文件名（默认为原文件名加前缀）')
    parser.add_argument('--start-row', type=int, default=2,
                       help='开始处理的行号（默认从第2行开始，跳过标题行）。\n'
                            '支持从指定行开始处理条码，便于断点续传或分批处理。')
    
    # 天聚数行API配置
    parser.add_argument('--tianapi-key', required=True, help='天聚数行API的密钥')
    
    return parser.parse_args()

# 条码格式验证函数已在上方定义，不再需要图片识别功能


def query_product_info_tianapi(barcode, tianapi_key):
    """
    使用天聚数行API查询商品信息
    
    Args:
        barcode: 条码数据
        tianapi_key: 天聚数行API密钥
    
    Returns:
        dict: 包含查询结果的字典
    """
    try:
        # 天聚数行API地址
        api_url = "https://apis.tianapi.com/barcode/index"
        
        # 构建请求参数
        params = {
            'key': tianapi_key,
            'barcode': barcode  # 注意：天聚数行使用barcode参数而不是code
        }
        
        # 发送API请求
        print(f"    正在查询商品信息: {barcode}")
        response = requests.get(api_url, params=params, timeout=10)
        response.raise_for_status()
        
        # 解析JSON响应
        result = response.json()
        
        # 检查API响应状态
        if result.get('code') == 200 and result.get('result'):
            # 查询成功，返回商品信息
            return {
                'success': True,
                'data': result['result']
            }
        else:
            # 查询失败，返回错误信息
            error_msg = result.get('msg', '未知错误')
            return {
                'success': False,
                'error': f"API返回错误: {error_msg}"
            }
            
    except requests.exceptions.RequestException as e:
        # 网络请求错误
        return {
            'success': False,
            'error': f"网络请求失败: {str(e)}"
        }
    except Exception as e:
        # 其他错误
        return {
            'success': False,
            'error': f"查询商品信息时发生错误: {str(e)}"
        }

def main():
    """主函数"""
    # 获取命令行参数
    args = parse_args()
    
    # 验证Excel文件是否存在
    if not os.path.exists(args.excel_file):
        print(f"错误: 文件 {args.excel_file} 不存在")
        return
    
    # 配置信息
    EXCEL_FILE = args.excel_file
    BARCODE_COLUMNS = args.barcode_cols
    TIANAPI_KEY = args.tianapi_key
    START_ROW = args.start_row  # 新增：起始行参数
    
    # 打印配置信息
    print(f"条码数字列: {BARCODE_COLUMNS}")
    print(f"起始处理行: {START_ROW} (从此行开始处理条码)")
    print(f"API地址: https://apis.tianapi.com/barcode/index")
    print(f"API密钥: {TIANAPI_KEY[:8]}...")
    print(f"实时保存: 是（每次查询后立即保存到文件，确保数据不丢失）")
    print(f"处理模式: 逐行处理，支持断点续传")
    
    # 设置输出文件名
    if args.output:
        output_file = args.output
    else:
        output_file = f"tianapi_条码查询结果_{os.path.basename(EXCEL_FILE)}"
    
    # 复制原始Excel文件
    print(f"创建Excel文件副本: {output_file}")
    shutil.copy2(EXCEL_FILE, output_file)
    
    # 加载工作簿
    print(f"正在读取Excel文件中的条码数据: {EXCEL_FILE}")
    wb = load_workbook(output_file)
    ws = wb.active
    
    # 检测Excel的最后一列位置
    max_col = ws.max_column
    start_col = max_col + 1  # 从最后一列的下一列开始写入
    
    # 定义天聚数行API返回的字段映射
    field_names = ['name', 'barcode', 'spec', 'brand', 'firm_name', 'firm_address', 
                  'firm_status', 'gross_weight', 'width', 'height', 'depth', 'goods_type', 'goods_pic']
    field_headers = ['商品名称', '条码', '规格', '品牌', '厂商名称', '厂商地址', 
                    '厂商状态', '毛重', '宽度', '高度', '深度', '商品类型', '商品图片']
    
    print(f"将从第 {start_col} 列开始写入 {len(field_names)} 个字段")
    
    # 写入列标题（第1行）
    for i, header in enumerate(field_headers):
        ws.cell(row=1, column=start_col + i, value=header)
    
    # 获取所有条码数据的位置信息
    barcode_positions = []
    for col_num in BARCODE_COLUMNS:
        for row_num in range(START_ROW, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value:  # 如果单元格有内容（条码数字）
                barcode_str = str(cell.value).strip()
                is_valid, clean_barcode = validate_barcode(barcode_str)
                if is_valid:
                    barcode_positions.append((row_num, col_num, clean_barcode))
    
    if not barcode_positions:
        print(f"在指定列 {BARCODE_COLUMNS} 中没有找到有效的条码数据")
        return
    
    print(f"\n正在处理条码查询...")
    print(f"找到 {len(barcode_positions)} 个有效条码需要处理（从第{START_ROW}行开始）")
    
    # 统计信息
    total_processed = 0
    success_count = 0
    
    # 处理每个条码
    for row_num, col_num, barcode_data in barcode_positions:
        try:
            total_processed += 1
            print(f"\n[{total_processed}/{len(barcode_positions)}] 处理第{row_num}行第{col_num}列的条码...")
            print(f"  条码数据: {barcode_data}")
            
            # 查询商品信息
            product_result = query_product_info_tianapi(barcode_data, TIANAPI_KEY)
            
            if product_result.get('success'):
                print(f"  查询成功: {product_result['data'].get('name', '未知商品')}")
                
                # 立即写入Excel文件
                # 实时写入功能：每次查询完成后立即保存，避免程序中断导致数据丢失
                data = product_result['data']
                for i, field_name in enumerate(field_names):
                    value = data.get(field_name, '')
                    ws.cell(row=row_num, column=start_col + i, value=value)
                
                # 立即保存文件，确保数据不丢失
                # 这是新增的实时保存功能，每处理一行就保存一次
                wb.save(output_file)
                print(f"  ✓ 已写入并保存到第{row_num}行")
                success_count += 1
            else:
                print(f"  查询失败: {product_result.get('error', '未知错误')}")
                # 写入错误信息，便于后续人工处理
                error_msg = product_result.get('error', '未知错误')
                ws.cell(row=row_num, column=start_col, value=f"错误: {error_msg}")
                for i in range(1, len(field_names)):
                    ws.cell(row=row_num, column=start_col + i, value='')
                # 立即保存文件（即使查询失败也要保存状态）
                wb.save(output_file)
                print(f"  - 已写入错误信息到第{row_num}行")
                
        except Exception as e:
            print(f"  ❌ 处理失败: {str(e)}")
            # 写入错误标记
            ws.cell(row=row_num, column=start_col, value=f"错误: 处理错误: {e}")
            for i in range(1, len(field_names)):
                ws.cell(row=row_num, column=start_col + i, value='')
            wb.save(output_file)
    
    # 关闭工作簿
    wb.close()
    
    # 显示最终统计信息
    print(f"\n=== 处理完成 ===")
    print(f"总处理数量: {total_processed}")
    print(f"成功查询: {success_count}")
    print(f"失败数量: {total_processed - success_count}")
    print(f"成功率: {success_count/total_processed*100:.1f}%" if total_processed > 0 else "成功率: 0%")
    print(f"输出文件: {output_file}")
    print(f"\n注意: 所有结果已实时保存到Excel文件中")

if __name__ == "__main__":
    main()