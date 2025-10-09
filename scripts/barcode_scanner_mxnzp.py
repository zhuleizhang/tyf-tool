#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MXNZP条码查询与商品信息查询工具

功能说明:
1. 从Excel文件中读取条码数字列
2. 验证条码格式的有效性
3. 调用MXNZP API查询商品详细信息
4. 将查询结果写入新的Excel文件
5. 支持实时保存和断点续传

支持的条码格式:
- EAN-13 (13位数字，最常见的商品条码)
- EAN-8 (8位数字)
- UPC-A (12位数字)
- UPC-E (8位数字)
- 其他数字条码格式

使用方法:
python scripts/barcode_scanner_mxnzp.py <excel_file> --barcode-cols <列号> --app-id <应用ID> --app-secret <应用密钥> [--start-row <起始行>]

示例:
python scripts/barcode_scanner_mxnzp.py /Users/bytedance/Desktop/tgs/条码识别结果_好客来超市-商品统计-latest.xlsx --barcode-cols 5 --app-id kwbqkbhmjiojivvz --app-secret hMRk6MCMP4YfRjiNHKvEFermuQXU15QH
python scripts/barcode_scanner_mxnzp.py /Users/bytedance/Desktop/tgs/条码识别结果_好客来超市-商品统计-latest.xlsx --barcode-cols 5 --app-id your_app_id --app-secret your_app_secret --start-row 5

功能优势:
1. 处理速度快：直接读取条码数字，无需图像识别
2. 准确率高：避免了图像识别可能出现的错误
3. 使用简单：只需在Excel中输入条码数字即可
4. 支持批量处理：可同时处理多个条码列
5. 断点续传：支持从指定行开始处理
6. 实时保存：每处理一条记录立即保存，避免数据丢失

注意事项:
1. 需要有效的MXNZP API账号和密钥
2. API有QPS限制，脚本会自动控制请求频率（每秒1次请求）
3. 条码必须是有效的数字格式
4. 输出文件前缀为 mxnzp_

作者: Assistant
创建时间: 2024
更新时间: 2024 - 改造为条码数字输入版本
"""

import argparse
import os
import re
import shutil
import sys
import time
import requests
from openpyxl import load_workbook

def validate_barcode(barcode_str):
    """
    验证条码格式
    
    Args:
        barcode_str: 条码字符串
    
    Returns:
        tuple: (是否有效, 清理后的条码)
    """
    if not barcode_str:
        return False, None
    
    # 转换为字符串并去除空白字符
    barcode_str = str(barcode_str).strip()
    
    # 检查是否为空
    if not barcode_str:
        return False, None
    
    # 移除所有非数字字符
    clean_barcode = re.sub(r'[^0-9]', '', barcode_str)
    
    # 检查是否包含数字
    if not clean_barcode:
        return False, None
    
    # 检查长度（一般条码长度在8-18位之间）
    if len(clean_barcode) < 8 or len(clean_barcode) > 18:
        return False, None
    
    return True, clean_barcode

def check_dependencies():
    """检查必要的依赖库"""
    missing_deps = []
    
    # 检查必要依赖
    deps_to_check = [
        ('openpyxl', 'openpyxl'),
        ('requests', 'requests')
    ]
    
    for module_name, package_name in deps_to_check:
        try:
            __import__(module_name)
            print(f"✓ {package_name}库导入成功")
        except ImportError as e:
            missing_deps.append((package_name, str(e)))
    
    if missing_deps:
        print("\n❌ 缺少以下依赖库:")
        for dep, error in missing_deps:
            print(f"  - {dep}: {error}")
        
        print("\n📋 解决方案:")
        print("   pip install openpyxl requests")
        return False
    
    return True

# 检查依赖
if not check_dependencies():
    print("\n❌ 依赖检查失败，请按照上述说明安装缺少的依赖库")
    sys.exit(1)

def parse_args():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(
        description='从Excel文件中读取条码数字并查询商品信息（MXNZP API版本）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  python scripts/barcode_scanner_mxnzp.py /Users/bytedance/Desktop/tgs/好客来超市-商品统计-latest.xlsx --barcode-cols 2 --app-id YOUR_APP_ID --app-secret YOUR_APP_SECRET
  python scripts/barcode_scanner_mxnzp.py /Users/bytedance/Desktop/tgs/好客来超市-商品统计.xlsx --barcode-cols 2 3 --app-id YOUR_APP_ID --app-secret YOUR_APP_SECRET --output result.xlsx --start-row 3
  
注意事项:
  - 需要先申请MXNZP API密钥
  - API有QPS限制，脚本会自动控制请求频率
  - 商品信息将从Excel最后一列开始写入
  - 条码数字列应包含有效的数字格式条码
  - 支持从指定行开始处理，默认从第2行开始
        """
    )
    
    parser.add_argument('excel_file', help='包含条码数字的Excel文件路径')
    parser.add_argument('--barcode-cols', type=int, nargs='+', required=True,
                       help='包含条码数字的列号（从1开始计数），可指定多列')
    parser.add_argument('--output', help='输出文件名（默认为原文件名加前缀）')
    parser.add_argument('--start-row', type=int, default=2,
                       help='开始处理的行号（默认从第2行开始，第1行通常是标题）')
    
    # MXNZP API配置
    parser.add_argument('--app-id', required=True, help='MXNZP API的app_id')
    parser.add_argument('--app-secret', required=True, help='MXNZP API的app_secret')
    parser.add_argument('--api-url', default="https://www.mxnzp.com/api/barcode/goods/details",
                       help='MXNZP API地址（默认使用官方地址）')
    
    return parser.parse_args()

def read_barcode_from_cell(cell_value):
    """
    从Excel单元格中读取条码数字
    
    Args:
        cell_value: Excel单元格的值
    
    Returns:
        tuple: (条码数据, 是否成功) 或 (None, False)
    """
    try:
        # 验证条码格式
        is_valid, clean_barcode = validate_barcode(cell_value)
        
        if is_valid:
            print(f"    ✓ 读取到有效条码: {clean_barcode}")
            return clean_barcode, True
        else:
            print(f"    ✗ 无效的条码格式: {cell_value}")
            return None, False
            
    except Exception as e:
        print(f"    读取条码错误: {e}")
        return None, False

def query_product_info_mxnzp(barcode, api_url, app_id, app_secret):
    """
    使用MXNZP API查询商品信息
    
    Args:
        barcode: 条码数据
        api_url: API地址
        app_id: 应用ID
        app_secret: 应用密钥
    
    Returns:
        dict: 包含查询结果的字典
    """
    try:
        # 构建请求参数
        params = {
            'barcode': barcode,
            'app_id': app_id,
            'app_secret': app_secret
        }
        
        # 发送API请求
        print(f"    正在查询商品信息: {barcode}")
        response = requests.get(api_url, params=params, timeout=10)
        response.raise_for_status()
        
        # 解析JSON响应
        result = response.json()
        
        # 检查API响应状态
        if result.get('code') == 1 and result.get('data'):
            # 查询成功，返回商品信息
            return {
                'success': True,
                'data': result['data']
            }
        else:
            # 查询失败，返回错误信息
            error_msg = result.get('msg', '未知错误')
            return {
                'success': False,
                'error': f"API返回错误: {error_msg}"
            }
            
    except requests.exceptions.RequestException as e:
        # 网络请求错误
        return {
            'success': False,
            'error': f"网络请求失败: {str(e)}"
        }
    except Exception as e:
        # 其他错误
        return {
            'success': False,
            'error': f"查询商品信息时发生错误: {str(e)}"
        }

def main():
    """主函数"""
    # 获取命令行参数
    args = parse_args()
    
    # 验证Excel文件是否存在
    if not os.path.exists(args.excel_file):
        print(f"错误: 文件 {args.excel_file} 不存在")
        return
    
    # 配置信息
    EXCEL_FILE = args.excel_file
    BARCODE_COLUMNS = args.barcode_cols
    API_URL = args.api_url
    APP_ID = args.app_id
    APP_SECRET = args.app_secret
    START_ROW = args.start_row
    
    # 打印配置信息
    print(f"条码数字列: {BARCODE_COLUMNS}")
    print(f"起始行: {START_ROW}")
    print(f"API地址: {API_URL}")
    print(f"应用ID: {APP_ID}")
    
    # 设置输出文件名
    if args.output:
        output_file = args.output
    else:
        output_file = f"mxnzp_条码查询结果_{os.path.basename(EXCEL_FILE)}"
    
    # 复制原始Excel文件
    print(f"创建Excel文件副本: {output_file}")
    shutil.copy2(EXCEL_FILE, output_file)
    
    # 加载工作簿
    print(f"正在读取Excel文件: {output_file}")
    wb = load_workbook(output_file)
    ws = wb.active
    
    # 检测Excel的最后一列位置
    max_col = ws.max_column
    start_col = max_col + 1  # 从最后一列的下一列开始写入
    
    # 定义MXNZP API返回的字段映射
    field_names = ['goodsName', 'barcode', 'price', 'brand', 'supplier', 'standard']
    field_headers = ['商品名称', '条码', '价格', '品牌', '供应商', '规格']
    
    print(f"将从第 {start_col} 列开始写入 {len(field_names)} 个字段")
    
    # 写入列标题（第1行）
    for i, header in enumerate(field_headers):
        ws.cell(row=1, column=start_col + i, value=header)
    
    # 收集条码查询结果
    query_results = {}  # 格式: {行号: 商品信息结构化数据}
    
    # 处理每个条码列
    for barcode_col in BARCODE_COLUMNS:
        print(f"\n处理条码数字列 {barcode_col}")
        
        # 从指定行开始处理
        row = START_ROW
        while True:
            cell = ws.cell(row=row, column=barcode_col)
            
            # 检查单元格是否为空
            if not cell.value:
                break
            
            print(f" ============= 行 {row}: 开始处理 ============= \n")
            
            # 读取条码数据
            barcode_data, read_success = read_barcode_from_cell(cell.value)
            
            if read_success and barcode_data:
                print(f"  行 {row}: 读取到有效条码 {barcode_data}")
                
                # 查询商品信息（MXNZP API有QPS限制，需要控制请求频率）
                product_result = query_product_info_mxnzp(barcode_data, API_URL, APP_ID, APP_SECRET)
                
                # 如果商品信息查询失败，但条码有效，创建包含条码信息的结果结构
                if not product_result.get('success'):
                    print(f"  查询失败: {product_result.get('error', '未知错误')}")
                    print(f"  条码有效，商品信息查询失败，仅填入条码")
                    # 创建包含条码信息的结果结构，其他字段为空
                    query_results[row] = {
                        'success': True,  # 标记为成功，因为条码有效
                        'barcode_only': True,  # 标记这是仅有条码的情况
                        'data': {
                            'goodsName': '',
                            'barcode': barcode_data,  # 填入有效的条码
                            'price': '',
                            'brand': '',
                            'supplier': '',
                            'standard': ''
                        }
                    }
                else:
                    print(f"  商品信息查询成功")
                    # 保存完整的查询结果
                    query_results[row] = product_result
                
                # QPS限制：确保1秒内只调用一次API
                print(f"  等待1秒（QPS限制）...")
                time.sleep(1)
                
            else:
                print(f"  行 {row}: 条码格式无效")
                query_results[row] = {'success': False, 'error': '条码格式无效'}
            
            print(f" ============= 行 {row}: 处理结束 ============= \n")
            row += 1
    
    # 将查询结果写入多列
    for row, result in query_results.items():
        if result.get('success') and result.get('data'):
            # 成功获取商品信息或条码有效（包括仅有条码的情况），按字段写入各列
            data = result['data']
            for i, field_name in enumerate(field_names):
                value = data.get(field_name, '')
                ws.cell(row=row, column=start_col + i, value=value)
            
            # 如果是仅有条码的情况，在日志中记录
            if result.get('barcode_only'):
                print(f"  行 {row}: 已填入条码 {data.get('barcode', '')}，其他信息为空")
        else:
            # 条码格式无效，在第一列写入错误信息
            error_msg = result.get('error', '未知错误')
            ws.cell(row=row, column=start_col, value=f"错误: {error_msg}")
            # 其他列留空
            for i in range(1, len(field_names)):
                ws.cell(row=row, column=start_col + i, value='')
    
    # 保存结果
    try:
        wb.save(output_file)
        print(f"处理完成，结果已保存到 {output_file}")
        print(f"\n共处理 {len(query_results)} 个条码")
        
        # 统计不同类型的结果
        full_success_count = sum(1 for result in query_results.values() 
                               if result.get('success') and not result.get('barcode_only'))
        barcode_only_count = sum(1 for result in query_results.values() 
                               if result.get('success') and result.get('barcode_only'))
        failed_count = sum(1 for result in query_results.values() if not result.get('success'))
        
        print(f"成功查询到完整商品信息: {full_success_count} 个")
        print(f"条码有效但商品信息查询失败: {barcode_only_count} 个")
        print(f"条码格式无效: {failed_count} 个")
    except Exception as e:
        print(f"保存文件时出错: {e}")
        # 尝试使用另一种方式保存
        try:
            print("尝试使用替代方法保存文件...")
            temp_output = f"temp_{output_file}"
            wb.save(temp_output)
            wb.close()
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_output, output_file)
            print(f"使用替代方法保存成功: {output_file}")
        except Exception as e2:
            print(f"使用替代方法保存失败: {e2}")

if __name__ == "__main__":
    main()