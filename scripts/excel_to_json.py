import argparse
import base64
import io
import json
import os
from openpyxl import load_workbook
from PIL import Image

'''
python scripts/excel_to_json.py /path/to/your/excel_file.xlsx

python scripts/excel_to_json.py /path/to/your/excel_file.xlsx --output result.json --indent 4 --sheet "Sheet2"
参数说明：

- --output ：指定输出的JSON文件路径
- --indent ：设置JSON文件的缩进空格数（默认为2）
- --sheet ：指定要处理的工作表名称（默认为活动工作表）

'''

# 解析命令行参数
def parse_args():
    parser = argparse.ArgumentParser(description="Excel转JSON工具（支持图片转Base64）")
    parser.add_argument("excel_file", help="Excel文件路径")
    parser.add_argument(
        "--output",
        help="输出的JSON文件路径，默认为'[原文件名].json'"
    )
    parser.add_argument(
        "--indent", 
        type=int, 
        default=2, 
        help="JSON缩进空格数，默认为2"
    )
    parser.add_argument(
        "--sheet", 
        help="要处理的工作表名称，默认为活动工作表"
    )
    return parser.parse_args()

def excel_to_json(excel_file, output_file=None, indent=2, sheet_name=None):
    # 验证Excel文件是否存在
    if not os.path.exists(excel_file):
        print(f"错误: 文件 {excel_file} 不存在")
        return False
    
    # 设置输出文件名
    if output_file is None:
        file_name = os.path.splitext(os.path.basename(excel_file))[0]
        output_file = f"{file_name}.json"
    
    # 加载工作簿
    print(f"正在读取Excel文件: {excel_file}")
    workbook = load_workbook(excel_file)
    
    # 选择工作表
    if sheet_name and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active
        if sheet_name:
            print(f"警告: 未找到工作表 '{sheet_name}'，使用活动工作表 '{sheet.title}'")
    
    print(f"处理工作表: {sheet.title}")
    
    # 获取所有图片的位置信息
    image_positions = []
    for idx, img in enumerate(sheet._images):
        row = img.anchor._from.row + 1  # 转换为1-indexed
        col = img.anchor._from.col + 1  # 转换为1-indexed
        image_positions.append((idx, row, col))
    
    # 获取表头（第一行作为键）
    headers = []
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value is not None:  # 只包含有值的列
            headers.append((col, str(cell_value)))
    
    if not headers:
        print("错误: 未在第一行找到有效的列名")
        return False
    
    print(f"找到 {len(headers)} 个有效列")
    
    # 处理每一行数据
    result = []
    for row_idx in range(2, sheet.max_row + 1):  # 从第二行开始（跳过表头）
        row_data = {}
        
        # 处理常规单元格数据
        for col_idx, header in headers:
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            
            # 检查该单元格是否有图片
            has_image = False
            image_idx = None
            for idx, row, col in image_positions:
                if row == row_idx and col == col_idx:
                    has_image = True
                    image_idx = idx
                    break
            
            # 如果是图片单元格，转换为base64
            if has_image and image_idx is not None:
                try:
                    # 获取图片数据并转换为base64
                    img = sheet._images[image_idx]
                    img_data = img._data()
                    pil_img = Image.open(io.BytesIO(img_data))
                    
                    # 确定图片格式
                    img_format = pil_img.format if pil_img.format else "PNG"
                    
                    buffered = io.BytesIO()
                    pil_img.save(buffered, format=img_format)
                    img_base64 = base64.b64encode(buffered.getvalue()).decode("utf-8")
                    
                    # 添加图片类型前缀
                    mime_type = f"image/{img_format.lower()}"
                    data_uri = f"data:{mime_type};base64,{img_base64}"
                    
                    row_data[header] = ''
                    # row_data[header] = data_uri
                    print(f"  行 {row_idx}, 列 '{header}': 已转换图片为base64")
                except Exception as e:
                    print(f"  处理行 {row_idx}, 列 '{header}' 的图片时出错: {e}")
                    row_data[header] = None
            else:
                # 非图片单元格，直接使用单元格值
                row_data[header] = cell_value
        
        # 只有当行数据不为空时才添加到结果中
        if any(value is not None for value in row_data.values()):
            result.append(row_data)
    
    # 关闭工作簿
    workbook.close()
    
    # 将结果包装在products字段中
    final_result = {"products": result}
    
    # 保存为JSON文件
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(final_result, f, ensure_ascii=False, indent=indent)
        print(f"转换完成，共 {len(result)} 行数据，结果已保存到 {output_file}")
        return True
    except Exception as e:
        print(f"保存JSON文件时出错: {e}")
        return False

def main():
    args = parse_args()
    excel_to_json(
        args.excel_file,
        args.output,
        args.indent,
        args.sheet
    )

if __name__ == "__main__":
    main()