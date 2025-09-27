import argparse
import base64
import io
import os
import shutil
import requests
import json
from openpyxl import load_workbook
from PIL import Image
from openpyxl.drawing.image import Image as XLImage

# 设置zbar库路径（macOS Homebrew）
if os.path.exists('/opt/homebrew/opt/zbar/lib'):
    zbar_lib_path = '/opt/homebrew/opt/zbar/lib'
    current_path = os.environ.get('DYLD_LIBRARY_PATH', '')
    if zbar_lib_path not in current_path:
        os.environ['DYLD_LIBRARY_PATH'] = f"{zbar_lib_path}:{current_path}"
try:
    from pyzbar import pyzbar
except ImportError:
    try:
        import pyzbar as pyzbar_module
        pyzbar = pyzbar_module
    except ImportError:
        print("错误: 请先安装pyzbar库: pip install pyzbar")
        exit(1)

# python scripts/barcode_scanner.py /Users/bytedance/Downloads/商品条码统计.xlsx
'''
python scripts/barcode_scanner.py /Users/bytedance/Downloads/商品条码统计.xlsx --image-cols 1 3 5

这个命令会同时处理：
- 第1列条码图片，商品信息结果从最后一列开始写入多个字段
- 第3列条码图片，商品信息结果从最后一列开始写入多个字段  
- 第5列条码图片，商品信息结果从最后一列开始写入多个字段
'''

# 解析命令行参数
def parse_args():
    parser = argparse.ArgumentParser(description="Excel条码图片商品信息查询工具")
    parser.add_argument("excel_file", help="Excel文件路径")
    parser.add_argument(
        "--image-cols", 
        type=int, 
        nargs="+",  # 允许输入多个值
        default=[1], 
        help="包含条码图片的列号列表(从1开始)，可指定多列，如 --image-cols 1 3 5"
    )
    # 移除result-cols参数，改为自动从最后一列开始写入
    parser.add_argument(
        "--api-provider",
        choices=["mxnzp", "tianapi"],
        # default="tianapi",
        default="mxnzp",
        help="选择API提供商: mxnzp (需要申请密钥) 或 tianapi (提供免费100次测试)"
    )
    parser.add_argument(
        "--api-url",
        help="自定义API地址（可选，会根据api-provider自动设置）"
    )
    parser.add_argument(
        "--app-id",
        help="API应用ID（mxnzp需要）"
    )
    parser.add_argument(
        "--app-secret",
        help="API应用密钥（mxnzp需要）"
    )
    parser.add_argument(
        "--tianapi-key",
        help="天聚数行API密钥（tianapi需要，免费注册获取100次测试额度）"
    )
    parser.add_argument(
        "--output",
        help="输出文件路径，默认为'条码查询结果_原文件名.xlsx'"
    )
    return parser.parse_args()

def decode_barcode_from_image(image_data):
    """从图片数据中识别条码"""
    try:
        # 将图片数据转换为PIL Image
        pil_img = Image.open(io.BytesIO(image_data))
        
        # 使用pyzbar识别条码
        barcodes = pyzbar.decode(pil_img)
        
        if barcodes:
            # 返回第一个识别到的条码数据
            barcode_data = barcodes[0].data.decode('utf-8')
            barcode_type = barcodes[0].type
            return barcode_data, barcode_type
        else:
            return None, None
    except Exception as e:
        print(f"条码识别错误: {e}")
        return None, None

def query_product_info_mxnzp(barcode, api_url, app_id, app_secret):
    """通过mxnzp API查询商品信息，返回结构化数据"""
    try:
        params = {
            'barcode': barcode,
            'app_id': app_id,
            'app_secret': app_secret
        }
        
        response = requests.get(api_url, params=params, timeout=10)
        response.raise_for_status()
        
        data = response.json()
        
        if data.get('code') == 1 and data.get('data'):
            product_data = data['data']
            # 返回mxnzp API的字段结构
            return {
                'success': True,
                'data': {
                    'goodsName': product_data.get('goodsName', ''),
                    'barcode': product_data.get('barcode', ''),
                    'price': product_data.get('price', ''),
                    'brand': product_data.get('brand', ''),
                    'supplier': product_data.get('supplier', ''),
                    'standard': product_data.get('standard', '')
                }
            }
        else:
            return {
                'success': False,
                'error': data.get('msg', '未知错误')
            }
            
    except requests.exceptions.RequestException as e:
        return {
            'success': False,
            'error': f"网络请求错误: {str(e)}"
        }
    except Exception as e:
        return {
            'success': False,
            'error': f"查询商品信息时发生错误: {str(e)}"
        }

def query_product_info_tianapi(barcode, tianapi_key):
    """通过天聚数行API查询商品信息，返回结构化数据"""
    try:
        api_url = "https://apis.tianapi.com/barcode/index"
        params = {
            'key': tianapi_key,
            'barcode': barcode
        }
        
        response = requests.get(api_url, params=params, timeout=10)
        response.raise_for_status()
        
        data = response.json()
        
        if data.get('code') == 200 and data.get('result'):
            result = data['result']
            # 返回tianapi API的字段结构
            return {
                'success': True,
                'data': {
                    'name': result.get('name', ''),
                    'barcode': result.get('barcode', ''),
                    'spec': result.get('spec', ''),
                    'brand': result.get('brand', ''),
                    'firm_name': result.get('firm_name', ''),
                    'firm_address': result.get('firm_address', ''),
                    'firm_status': result.get('firm_status', ''),
                    'gross_weight': result.get('gross_weight', ''),
                    'width': result.get('width', ''),
                    'height': result.get('height', ''),
                    'depth': result.get('depth', ''),
                    'goods_type': result.get('goods_type', ''),
                    'goods_pic': result.get('goods_pic', '')
                }
            }
        elif data.get('code') == 130:
            return {
                'success': False,
                'error': "今日免费次数已用完，请明天再试或升级套餐"
            }
        elif data.get('code') == 140:
            return {
                'success': False,
                'error': "API密钥无效，请检查tianapi_key参数"
            }
        else:
            return {
                'success': False,
                'error': data.get('msg', '数据返回为空')
            }
            
    except requests.exceptions.RequestException as e:
        return {
            'success': False,
            'error': f"网络请求错误: {str(e)}"
        }
    except Exception as e:
        return {
            'success': False,
            'error': f"查询商品信息时发生错误: {str(e)}"
        }

def query_product_info(barcode, api_provider, **kwargs):
    """根据API提供商查询商品信息，返回结构化数据"""
    if api_provider == "mxnzp":
        return query_product_info_mxnzp(
            barcode, 
            kwargs.get('api_url'), 
            kwargs.get('app_id'), 
            kwargs.get('app_secret')
        )
    elif api_provider == "tianapi":
        return query_product_info_tianapi(
            barcode, 
            kwargs.get('tianapi_key')
        )
    else:
        return {
            'success': False,
            'error': f"不支持的API提供商: {api_provider}"
        }

def main():
    # 获取命令行参数
    args = parse_args()
    
    # 验证Excel文件是否存在
    if not os.path.exists(args.excel_file):
        print(f"错误: 文件 {args.excel_file} 不存在")
        return
    
    # 移除图片列和结果列数量匹配验证（现在自动从最后一列开始写入）
    
    # 根据API提供商设置默认配置
    api_provider = args.api_provider
    
    if api_provider == "mxnzp":
        # 验证mxnzp必需参数
        if not args.app_id or not args.app_secret:
            print("错误: 使用mxnzp API需要提供 --app-id 和 --app-secret 参数")
            print("\n获取mxnzp API密钥:")
            print("1. 访问 https://www.mxnzp.com/")
            print("2. 注册账号并申请API密钥")
            print("3. 使用命令: python scripts/barcode_scanner.py [文件] --api-provider mxnzp --app-id [你的ID] --app-secret [你的密钥]")
            return
        
        api_url = args.api_url or "https://www.mxnzp.com/api/barcode/goods/details"
        api_config = {
            'api_url': api_url,
            'app_id': args.app_id,
            'app_secret': args.app_secret
        }
        
    elif api_provider == "tianapi":
        # 验证tianapi必需参数
        if not args.tianapi_key:
            print("错误: 使用tianapi需要提供 --tianapi-key 参数")
            print("\n获取天聚数行API密钥:")
            print("1. 访问 https://www.tianapi.com/")
            print("2. 注册账号并申请条码查询API（免费100次测试）")
            print("3. 使用命令: python scripts/barcode_scanner.py [文件] --api-provider tianapi --tianapi-key [你的密钥]")
            return
        
        api_config = {
            'tianapi_key': args.tianapi_key
        }
    
    # 配置信息
    EXCEL_FILE = args.excel_file
    IMAGE_COLUMNS = args.image_cols
    
    # 打印配置信息
    print(f"条码图片列: {IMAGE_COLUMNS}")
    print(f"API提供商: {api_provider}")
    if api_provider == "mxnzp":
        print(f"API地址: {api_config['api_url']}")
    elif api_provider == "tianapi":
        print(f"API地址: https://apis.tianapi.com/barcode/index")
    
    # 设置输出文件名
    if args.output:
        output_file = args.output
    else:
        output_file = f"条码查询结果_{os.path.basename(EXCEL_FILE)}"
    
    # 复制原始Excel文件
    print(f"创建Excel文件副本: {output_file}")
    shutil.copy2(EXCEL_FILE, output_file)
    
    # 加载原始工作簿（用于读取图片数据）
    print(f"正在读取原始Excel文件中的条码图片: {EXCEL_FILE}")
    source_wb = load_workbook(EXCEL_FILE)
    source_sheet = source_wb.active
    
    # 检测Excel的最后一列位置
    max_col = source_sheet.max_column
    start_col = max_col + 1  # 从最后一列的下一列开始写入
    
    # 定义字段映射
    if api_provider == "mxnzp":
        field_names = ['goodsName', 'barcode', 'price', 'brand', 'supplier', 'standard']
        field_headers = ['商品名称', '条码', '价格', '品牌', '供应商', '规格']
    elif api_provider == "tianapi":
        field_names = ['name', 'barcode', 'spec', 'brand', 'firm_name', 'firm_address', 
                      'firm_status', 'gross_weight', 'width', 'height', 'depth', 'goods_type', 'goods_pic']
        field_headers = ['商品名称', '条码', '规格', '品牌', '厂商名称', '厂商地址', 
                        '厂商状态', '毛重', '宽度', '高度', '深度', '商品类型', '商品图片']
    
    print(f"将从第 {start_col} 列开始写入 {len(field_names)} 个字段")
    
    # 收集条码识别和商品查询结果
    query_results = {}  # 格式: {行号: 商品信息结构化数据}
    image_positions = []
    
    # 获取所有图片的位置信息
    for idx, img in enumerate(source_sheet._images):
        row = img.anchor._from.row + 1  # 转换为1-indexed
        col = img.anchor._from.col + 1  # 转换为1-indexed
        image_positions.append((idx, row, col))
    
    # 处理每个图片列
    for image_col in IMAGE_COLUMNS:
        print(f"\n处理条码图片列 {image_col}")
        
        # 找出当前图片列中的所有图片
        column_images = [(idx, row, col) for idx, row, col in image_positions if col == image_col]
        
        if not column_images:
            print(f"警告: 列 {image_col} 中未找到图片")
            continue
        
        # 处理当前列的每个图片
        for idx, row, col in column_images:
            try:
                # 获取图片数据
                img = source_sheet._images[idx]
                img_data = img._data()
                
                # 识别条码
                barcode_data, barcode_type = decode_barcode_from_image(img_data)
                
                if barcode_data:
                    print(f"  行 {row}: 识别到条码 {barcode_data} (类型: {barcode_type})")
                    
                    # 查询商品信息
                    product_result = query_product_info(barcode_data, api_provider, **api_config)
                    
                    # 保存查询结果
                    query_results[row] = product_result
                    
                    if product_result.get('success'):
                        print(f"  商品信息查询成功")
                    else:
                        print(f"  查询失败: {product_result.get('error', '未知错误')}")
                else:
                    print(f"  行 {row}: 未识别到条码")
                    query_results[row] = {'success': False, 'error': '未识别到条码'}
                
            except Exception as e:
                print(f"  处理行 {row} 图片时出错: {e}")
                query_results[row] = {'success': False, 'error': f'处理错误: {e}'}
    
    # 关闭原始工作簿
    source_wb.close()
    
    # 以只读方式加载复制后的工作簿（只修改单元格值，不处理图片）
    print(f"\n正在将商品信息写入到: {output_file}")
    target_wb = load_workbook(output_file, read_only=False, keep_vba=True, data_only=False, keep_links=True)
    target_sheet = target_wb.active
    
    # 写入列标题（第1行）
    for i, header in enumerate(field_headers):
        target_sheet.cell(row=1, column=start_col + i, value=header)
    
    # 将查询结果写入多列
    for row, result in query_results.items():
        if result.get('success') and result.get('data'):
            # 成功获取商品信息，按字段写入各列
            data = result['data']
            for i, field_name in enumerate(field_names):
                value = data.get(field_name, '')
                target_sheet.cell(row=row, column=start_col + i, value=value)
        else:
            # 查询失败，在第一列写入错误信息
            error_msg = result.get('error', '未知错误')
            target_sheet.cell(row=row, column=start_col, value=f"错误: {error_msg}")
            # 其他列留空
            for i in range(1, len(field_names)):
                target_sheet.cell(row=row, column=start_col + i, value='')
    
    # 保存结果
    try:
        target_wb.save(output_file)
        print(f"处理完成，结果已保存到 {output_file}")
    except Exception as e:
        print(f"保存文件时出错: {e}")
        # 尝试使用另一种方式保存
        try:
            print("尝试使用替代方法保存文件...")
            # 创建临时文件名
            temp_output = f"temp_{output_file}"
            target_wb.save(temp_output)
            # 关闭工作簿
            target_wb.close()
            # 替换原文件
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_output, output_file)
            print(f"使用替代方法保存成功: {output_file}")
        except Exception as e2:
            print(f"使用替代方法保存失败: {e2}")

if __name__ == "__main__":
    main()