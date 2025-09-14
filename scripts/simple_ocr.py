import argparse
import base64
import io
import os
import shutil
import requests
from openpyxl import load_workbook
from PIL import Image
from openpyxl.drawing.image import Image as XLImage

# python simple_ocr.py /Users/bytedance/Downloads/好客来超市-商品统计.xlsx
'''
python simple_ocr.py /Users/bytedance/Downloads/好客来超市-商品统计.xlsx --image-cols 1 3 5 --result-cols 2 4 6

这个命令会同时处理：
- 第1列图片，结果写入第2列
- 第3列图片，结果写入第4列
- 第5列图片，结果写入第6列
'''

# 解析命令行参数
def parse_args():
    parser = argparse.ArgumentParser(description="Excel图片OCR文字识别工具")
    parser.add_argument("excel_file", help="Excel文件路径")
    parser.add_argument(
        "--image-cols", 
        type=int, 
        nargs="+",  # 允许输入多个值
        default=[1], 
        help="包含图片的列号列表(从1开始)，可指定多列，如 --image-cols 1 3 5"
    )
    parser.add_argument(
        "--result-cols", 
        type=int, 
        nargs="+",  # 允许输入多个值
        default=[2], 
        help="写入识别结果的列号列表(从1开始)，必须与image-cols数量相同，如 --result-cols 2 4 6"
    )
    parser.add_argument(
        "--api-url",
        default="http://0.0.0.0:8000/api/v1/ocr/recognize",
        help="OCR服务API地址"
    )
    parser.add_argument(
        "--output",
        help="输出文件路径，默认为'OCR处理结果_原文件名.xlsx'"
    )
    return parser.parse_args()

def main():
    # 获取命令行参数
    args = parse_args()
    
    # 验证Excel文件是否存在
    if not os.path.exists(args.excel_file):
        print(f"错误: 文件 {args.excel_file} 不存在")
        return
    
    # 验证图片列和结果列数量是否匹配
    if len(args.image_cols) != len(args.result_cols):
        print(f"错误: 图片列数量({len(args.image_cols)})与结果列数量({len(args.result_cols)})不匹配")
        return
    
    # 配置信息
    EXCEL_FILE = args.excel_file
    IMAGE_COLUMNS = args.image_cols
    RESULT_COLUMNS = args.result_cols
    API_URL = args.api_url
    
    # 打印配置信息
    print(f"图片列: {IMAGE_COLUMNS}")
    print(f"结果列: {RESULT_COLUMNS}")
    
    # 设置输出文件名
    if args.output:
        output_file = args.output
    else:
        output_file = f"OCR处理结果_{os.path.basename(EXCEL_FILE)}"
    
    # 复制原始Excel文件
    print(f"创建Excel文件副本: {output_file}")
    shutil.copy2(EXCEL_FILE, output_file)
    
    # 加载原始工作簿（用于读取图片数据）
    print(f"正在读取原始Excel文件中的图片: {EXCEL_FILE}")
    source_wb = load_workbook(EXCEL_FILE)
    source_sheet = source_wb.active
    
    # 收集图片数据和OCR结果
    ocr_results = {}  # 格式: {(行号, 结果列号): 识别文本}
    image_positions = []
    
    # 获取所有图片的位置信息
    for idx, img in enumerate(source_sheet._images):
        row = img.anchor._from.row + 1  # 转换为1-indexed
        col = img.anchor._from.col + 1  # 转换为1-indexed
        image_positions.append((idx, row, col))
    
    # 处理每个图片列和对应的结果列
    for i, image_col in enumerate(IMAGE_COLUMNS):
        result_col = RESULT_COLUMNS[i]
        print(f"\n处理图片列 {image_col} -> 结果列 {result_col}")
        
        # 找出当前图片列中的所有图片
        column_images = [(idx, row, col) for idx, row, col in image_positions if col == image_col]
        
        if not column_images:
            print(f"警告: 列 {image_col} 中未找到图片")
            continue
        
        # 处理当前列的每个图片
        for idx, row, col in column_images:
            try:
                # 获取图片数据并转换为base64
                img = source_sheet._images[idx]
                img_data = img._data()
                pil_img = Image.open(io.BytesIO(img_data))
                buffered = io.BytesIO()
                pil_img.save(buffered, format="PNG")
                img_base64 = base64.b64encode(buffered.getvalue()).decode("utf-8")
                
                # 调用OCR服务
                headers = {"Content-Type": "application/json"}
                payload = {"image_base64": img_base64}
                response = requests.post(API_URL, headers=headers, json=payload)
                response.raise_for_status()
                result = response.json()
                
                # 解析OCR结果
                recognized_text = ""
                if result.get("code") == 0 and "data" in result:
                    recognized_text = result["data"].get("text", "")
                else:
                    print(f"OCR服务返回错误: {result.get('msg', '未知错误')}")
                
                # 保存OCR结果（键为行号和结果列号的元组）
                ocr_results[(row, result_col)] = recognized_text
                print(f"  已处理行 {row}: {recognized_text}")
                
            except Exception as e:
                print(f"  处理行 {row} 图片时出错: {e}")
    
    # 关闭原始工作簿
    source_wb.close()
    
    # 以只读方式加载复制后的工作簿（只修改单元格值，不处理图片）
    print(f"\n正在将OCR结果写入到: {output_file}")
    target_wb = load_workbook(output_file, read_only=False, keep_vba=True, data_only=False, keep_links=True)
    target_sheet = target_wb.active
    
    # 将OCR结果写入指定列
    for (row, col), text in ocr_results.items():
        target_sheet.cell(row=row, column=col, value=text)
    
    # 保存结果
    try:
        target_wb.save(output_file)
        print(f"处理完成，结果已保存到 {output_file}")
    except Exception as e:
        print(f"保存文件时出错: {e}")
        # 尝试使用另一种方式保存
        try:
            print("尝试使用替代方法保存文件...")
            # 创建临时文件名
            temp_output = f"temp_{output_file}"
            target_wb.save(temp_output)
            # 关闭工作簿
            target_wb.close()
            # 替换原文件
            if os.path.exists(output_file):
                os.remove(output_file)
            os.rename(temp_output, output_file)
            print(f"使用替代方法保存成功: {output_file}")
        except Exception as e2:
            print(f"使用替代方法保存失败: {e2}")

if __name__ == "__main__":
    main()